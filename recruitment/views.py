import joblib
import json
import openpyxl
import os
import pandas as pd
from .models import Pelamar, Seleksi, Departemen, Perteker
from .forms import *
from sklearn.naive_bayes  import CategoricalNB
from sklearn.preprocessing import LabelEncoder
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.styles import Font, Alignment

# Create your views here.
def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('recruitment:index')
            else:
                # messages.error(request, 'Username atau password salah.')
                context = {
                    'form': form,
                    'error': 'Invalid Username or Password',
                    'title': 'GTI',
                    'subtitle': 'Recruitment',
                    'subsubtitle': 'System',
                }
                return render(request, 'recruitment/login.html', context)
        else:
            # messages.error(request, 'Form tidak valid. Coba lagi.')
            context = {
                'form': form,
                'error': 'Invalid Username or Password',
                'title': 'GTI',
                'subtitle': 'Rekrutmen',
                'subsubtitle': 'System',
            }
            return render(request, 'recruitment/login.html', context)
    
    # GET request
    form = AuthenticationForm()
    context = {
        'title': 'GTI',
        'subtitle': 'Recruitmen',
        'subsubtitle': 'System',
        'form': form,
    }
    return render(request, 'recruitment/login.html', context)

def user_logout(request):
    logout(request)
    return redirect('recruitment:login')

# =--------------------------------------------------
@login_required
def index(request):
    dept = Departemen.objects.all()
    dept_count = dept.count()
    pelamar = Pelamar.objects.all()
    pelamar_count = pelamar.count()
    perteker = Perteker.objects.all()
    perteker_count = perteker.count()
    user = CustomUser.objects.count()
    context = {
        'title'         : 'Dashboard',
        'subtitle'      : 'HR-Recruitment & Assesment',
        'daftar'        : '',
        'dept'          : dept_count,
        'pelamar'       : pelamar_count,
        'ptk'           : perteker_count,
        'user'          : user,
        'section'       : 'dashboard',
    }
    return render(request, 'recruitment/index.html', context)
# ---------------------------------------------------------------------------
@login_required
def listDepartemen(request):
    departemen = Departemen.objects.all()

    context = {
        'title'     : 'Departemen',
        'daftar'    : 'Daftar',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'departemens'     : departemen
    }

    return render(request, 'recruitment/departemen/list.html', context)

@login_required
def createDepartemen(request):
    form = DepartemenForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
        return redirect('recruitment:list_departemen')
    context = {
        'title'     : 'Tambah Departemen',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'forms'     : form
    }
    return render(request, 'recruitment/departemen/create.html', context)

@login_required
def updateDepartemen(request, update_id):
    update = Departemen.objects.get(id=update_id)
    data = {
        'nama'  : update.nama,
    }
    form = DepartemenForm(request.POST or None, initial=data, instance=update)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
        return redirect('recruitment:list_departemen')

    context = {
        'title'     : 'Update Departemen',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'forms'     : form,
    }

    return render(request, 'recruitment/departemen/create.html', context)

@login_required
def deleteDepartemen(request, delete_id):
    Departemen.objects.filter(id = delete_id).delete()
    return redirect('recruitmens:list_departemen')

@login_required
def departemenExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Departemen"

    # Judul besar di baris 1
    ws.merge_cells('A1:B1')  # gabungkan dari kolom A sampai F
    ws['A1'] = "Data Departemen"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['ID', 'Nama']
    ws.append(headers)

    # Data
    for s in Departemen.objects.all():
        ws.append([
            s.id,
            s.nama,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departemen.xlsx'
    wb.save(response)
    return response

@login_required
def filter_departements(request):
    query = request.GET.get('q','')
    print("receive query:", query)

    if query:
        departemens = Departemen.objects.filter(
            Q(departemen__icontains=query)
        )
    else:
        departemens = Departemen.objects.all()
    
    departemen_list = []
    for departemen in departemens:

        departemen_list.append({
            'id':departemen.id,
            'dept_name':departemen.departemen,
        })

    return JsonResponse({'departemen':departemen_list})
# --------------------------------------------------------------------------------------------------------------
@login_required
def listPelamar(request):
    pelamar = Pelamar.objects.all().order_by('-tanggal')
    
    context = {
        'title'     : 'Pelamar',
        'daftar'    : 'Daftar',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'pelamars'  : pelamar
    }

    return render(request, 'recruitment/pelamar/list.html', context)

@login_required
def createPelamar(request):
    form = PelamarForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Data berhasil disimpan.')
            return redirect('recruitment:list_pelamar')
        else:
            print("Form errors:", form.errors)
            messages.error(request, 'Form tidak valid. Periksa kembali input Anda.')

    context = {
        'title'     : 'Input Pelamar',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'forms'     : form
    }
    return render(request, 'recruitment/pelamar/create.html', context)

@login_required
def updatePelamar(request, update_id):
    update = Pelamar.objects.get(id=update_id)
    data = {
        'nama'      : update.nama,
        'gender'    : update.gender,
        'usia'      : update.usia,
        'pendidikan': update.pendidikan,
        'jurusan'   : update.jurusan,
        'alamamater': update.almamater,
        'pengalaman': update.pengalaman,
        'keahlian'  : update.keahlian,
        'pertaker'  : update.perteker,
        'alamat'    : update.alamat,
        'phone'     : update.phone,
    }
    form = PelamarForm(request.POST or None, initial=data, instance=update)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
        return redirect('recruitment:list_pelamar')

    context = {
        'title'     : 'Update Pelamar',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'forms'     : form,
    }

    return render(request, 'recruitment/pelamar/create.html', context)

@login_required
def deletePelamar(request, delete_id):
    Pelamar.objects.filter(id = delete_id).delete()
    return redirect('recruitment:list_pelamar')

@login_required
def pelamarExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Pelamar"

    # Judul besar di baris 1
    ws.merge_cells('A1:L1')  # gabungkan dari kolom A sampai F
    ws['A1'] = "Data Pelamar"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['ID', 'Nama', 'Gender', 'Usia', 'Pendidikan', 'Jurusan', 'Almamater', 'Pengalaman', 'Keahlian', 'Perteker', 'Alamat', 'Phone']
    ws.append(headers)

    # Data
    for s in Pelamar.objects.all():
        ws.append([
            s.id,
            s.nama,
            s.gender,
            s.usia,
            s.pendidikan,
            s.jurusan,
            s.almamater,
            s.pengalaman,
            s.keahlian,
            s.perteker.open_poss,
            s.alamat,
            s.phone,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pelamar.xlsx'
    wb.save(response)
    return response
# ---------------------------------------------------------------------------------------------
@login_required
def listPerteker(request):
    user = request.user
    if user.groups.filter(name='Admin').exists():
        # Jika user berada di grup admin, tampilkan semua data
        perteker = Perteker.objects.all()
    else:
        # Jika bukan admin, tampilkan data perteker dari user yang satu grup
        user_groups = user.groups.all()
        users_in_same_group = CustomUser.objects.filter(groups__in=user_groups).distinct()
        perteker = Perteker.objects.filter(user__in=users_in_same_group)

    # perteker = Perteker.objects.all()
    
    context = {
        'title'     : 'PTK',
        'daftar'    : 'Daftar',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'pertekers' : perteker
    }

    return render(request, 'recruitment/perteker/list.html', context)
 


@login_required
def createPerteker(request):
    # form = PertekerForm(request.POST or None)
    # if request.method == 'POST':
    #     if form.is_valid():
    #         form.save()
    #     return redirect('recruitment:list_perteker')

    form = PertekerForm(request.POST or None, user = request.user)
    
    if request.method == 'POST':
        if form.is_valid():
            print("Current user:", request.user)
            print("User type:", type(request.user))
            ptk_instance = form.save(commit=False)
            ptk_instance.user = request.user
            ptk_instance.save()
            messages.success(request, 'Data berhasil disimpan.')
            return redirect('recruitment:list_perteker')
        else:
            print("Form errors:", form.errors)
            messages.error(request, 'Form tidak valid. Periksa kembali input Anda.')
    context = {
        'title'     : 'Permintaan Tenaga Kerja',
        'daftar'    : 'Daftar',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'forms'     : form
    }
    return render(request, 'recruitment/perteker/create.html', context)

@login_required
def updatePerteker(request, update_id):
    update = Perteker.objects.get(id=update_id)
    data = {
        'gender'        : update.gender,
        'open_poss'     : update.open_poss,
        'batas_usia'    : update.batas_usia,
        'pendidikan_min': update.pendidikan_min,
        'jurusan'       : update.jurusan,
        'pengalaman'    : update.pengalaman,
        'jumlah'        : update.jumlah,
    }
    form = PertekerForm(request.POST or None, initial=data, instance=update)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
        return redirect('recruitment:list_perteker')

    context = {
        'title'     : 'Update Permintaan Tenaga Kerja',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'forms'     : form,
    }

    return render(request, 'recruitment/perteker/create.html', context)

@login_required
def deletePerteker(request, delete_id):
    Perteker.objects.filter(id = delete_id).delete()
    return redirect('recruitmen:list_perteker')

@login_required
def pertekerExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data PTK"

    # Judul besar di baris 1
    ws.merge_cells('A1:H1')  # gabungkan dari kolom A sampai F
    ws['A1'] = "Data PTK"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['ID', 'Gender', 'Posisi', 'Batas Usia', 'Pendidikan', 'Jurusan', 'Pengalaman', 'Jumlah']
    ws.append(headers)

    # Data
    for s in Perteker.objects.all():
        ws.append([
            s.id,
            s.gender,
            s.open_poss,
            s.batas_usia,
            s.pendidikan_min,
            s.jurusan,
            s.pengalaman,
            s.jumlah,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ptk.xlsx'
    wb.save(response)
    return response
    # ----------------------------------------------------------------------------------------------

def preprocess_data(df):
    # Kategorisasi Pendidikan
    def kategori_pendidikan(p):
        if p in ['SD/MI', 'SMP/MTS', 'SMA/SMK/MA']:
            return 'Umum'
        else:
            return 'Tinggi'
    
    # Kategorisasi Usia
    def kategori_usia(u):
        return 'Muda' if u < 35 else 'Dewasa'
    
    df['Pendidikan'] = df['Pendidikan'].apply(kategori_pendidikan)
    df['Usia'] = df['Usia'].apply(kategori_usia)

    return df

@login_required
def listSeleksi(request):
    # load model dan encoder yang sudah dilatih
    model_path = os.path.join(settings.BASE_DIR, 'recruitment/static', 'model_cakar.joblib')
    encoders_path = os.path.join(settings.BASE_DIR, 'recruitment/static', 'encoders_cakar.joblib')
    le_target_path = os.path.join(settings.BASE_DIR, 'recruitment/static', 'le_target_cakar.joblib')

    model = joblib.load(model_path)
    encoders = joblib.load(encoders_path)
    le_target = joblib.load(le_target_path)


    user = request.user

    if user.groups.filter(name='Admin').exists():
        seleksi_qs = Seleksi.objects.all().order_by('-pelamar')
    else:
        seleksi_qs = Seleksi.objects.filter(pelamar__perteker__user=request.user).order_by('-tanggal')
    #     user_groups = user.groups.all()

    #     # Ambil semua user dalam grup yang sama
    #     users_in_same_group = CustomUser.objects.filter(groups__in=user_groups).distinct()

    #     # Ambil pelamar yang perteker-nya dibuat oleh user dalam grup itu
    #     pelamars_in_group = Pelamar.objects.filter(perteker__user__in=users_in_same_group)

    #     # Ambil seleksi berdasarkan pelamar tersebut
    #     seleksi_qs = Seleksi.objects.filter(pelamar__in=pelamars_in_group)

    # seleksi_qs = Seleksi.objects.filter(pelamar__perteker__user=request.user)
    # penampungan data
    all_data = []

    # load data dari database
    for obj in seleksi_qs:
        record = {
            'id'        : obj.id,
            'poss'      : obj.pelamar.perteker.open_poss,
            'Nama'      : obj.pelamar.nama,
            'Gender'    : obj.pelamar.gender,
            'Pendidikan': obj.pelamar.pendidikan,
            'Pengalaman': obj.pelamar.pengalaman,
            'Usia'      : obj.pelamar.usia,
            'Psikotest' : obj.nilai_psikotest,
            'Interview' : obj.nilai_interview,
            'Status'    : obj.status,
            'Prediksi'  : 'Belum lengkap',
            'Catatan'   : obj.catatan,
            'Tanggal'   : obj.tanggal
        }

        if all([obj.pelamar.gender, obj.pelamar.pendidikan, obj.pelamar.pengalaman, obj.pelamar.usia, obj.nilai_psikotest, obj.nilai_interview]):
            # lakukan prediksi untuk data yang sudah lengkap
            # record = normalisasi_nilai(record)
            df_row  = pd.DataFrame([record]).copy()
            # preprocessing data
            df_row  = preprocess_data(df_row)
            fitur   = ['Gender', 'Pendidikan', 'Pengalaman', 'Usia', 'Psikotest', 'Interview']
            for col in fitur:
                df_row[col] = encoders[col].transform(df_row[col])
            X   = df_row[fitur]
            pred = model.predict(X)
            hasil = le_target.inverse_transform(pred)
            record['Prediksi'] = hasil[0]
        
        all_data.append(record)

    context = {
        'title'     : 'Monitoring Seleksi',
        'daftar'    : 'Daftar',
        'subtitle'  : 'HR-Recruitment & Assesment',
        'seleksis'  : all_data,
    }
    return render(request, 'recruitment/seleksi/list.html', context)

@login_required
def updateSeleksi(request, update_id):
    seleksi = get_object_or_404(Seleksi, id = update_id)
    if request.method == 'POST':
        seleksi.nilai_psikotest = request.POST.get('psikotest')
        seleksi.nilai_interview = request.POST.get('interview')
        seleksi.status          = request.POST.get('status')
        seleksi.catatan         = request.POST.get('catatan')
        seleksi.save()
        return redirect('recruitment:list_seleksi')

    context = {
        'title'     : 'Update Seleksi',
        'subtitle'  : 'HR Recruitment & Assesment',
        'seleksi'   : seleksi,
        'pil_nilai' : ['Baik', 'Cukup', 'Kurang'],
        'pil_status': ['Diterima', 'Ditolak', 'Mundur']
    }

    return render(request, 'recruitment/seleksi/update.html', context)

@login_required
def seleksiExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Monitoring Seleksi"

    # Judul besar di baris 1
    ws.merge_cells('A1:I1')  # gabungkan dari kolom A sampai F
    ws['A1'] = "Data Monitoring Seleksi"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['ID', 'Nama', 'Gender', 'Pendidikan', 'Pengalaman',  'Usia', 'Psikotest', 'Interview', 'Status']
    ws.append(headers)

    # Data
    for s in Seleksi.objects.all():
        ws.append([
            s.id,
            s.pelamar.nama,
            s.pelamar.gender,
            s.pelamar.pendidikan,
            s.pelamar.pengalaman,
            s.pelamar.usia,
            s.nilai_psikotest,
            s.nilai_interview,
            s.status,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Monitoring.xlsx'
    wb.save(response)
    return response

# ?-------------------------------------
@login_required
def list_user(request):
    user = CustomUser.objects.filter(is_staff=False)

    for users in user:
        users.role = ",".join([group.name for group in users.groups.all()])
        # print(users.role)
    context = {
        'title' : 'User',
        'daftar' : 'Daftar',
        'subtitle'  : 'HR Recruitment & Assesment',
        'users' : user,
    }

    return render(request, 'recruitment/user/list.html', context)

@login_required
def create_user(request):
    form = CustomUserCreationForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password1'])
            user.save()
        else:
            print(form.errors)
        return redirect('recruitment:list_user')

    context = {
        'title'     :'User',
        'subtitle'  :'Tambah',
        'forms'     : form,
    }
    
    return render(request, 'recruitment/user/create.html', context)

@login_required
def update_user(request, update_id):
    user_update = CustomUser.objects.get(id=update_id)
    data = {
        'username'      : user_update.username,
        'email'         : user_update.email,
        'first_name'    : user_update.first_name,
        'last_name'     : user_update.last_name,
        'departemen'    : user_update.departemen,
    }
    form_user = CustomUserUpdateForm(request.POST or None, initial=data, instance = user_update)
    
    if request.method == 'POST':
        if form_user.is_valid():
            form_user.save()
            
        return redirect('recruitment:list_user')
    context = {
        'title':'User',
        'subtitle':'Update',
        'forms':form_user,
        
    }
    return render(request, 'recruitment/user/create.html', context)

@login_required
def delete_user(request, delete_id):
    CustomUser.objects.filter(id=delete_id).delete()
    return redirect('recruitment:list_user')   
@login_required
def userExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data User"

    # Judul besar di baris 1
    ws.merge_cells('A1:C1')  # gabungkan dari kolom A sampai F
    ws['A1'] = "Data User"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['ID', 'Username', 'Email', 'Nama Depan', 'Nama Belakang']
    ws.append(headers)

    # Data
    for s in CustomUser.objects.all():
        ws.append([
            s.id,
            s.username,
            s.email,
            s.first_name,
            s.last_name,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=user.xlsx'
    wb.save(response)
    return response